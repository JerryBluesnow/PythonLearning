import sys
import os
import datetime as dt
import json
import tushare as ts
from openpyxl import load_workbook

columns_first_day = [12, 13, 14, 15]
columns_mid_1_day = [17, 18, 19, 20]
columns_mid_2_day = [22, 23, 24, 25]
columns_last_day = [27, 28, 29, 30]
setup_base_line = 10000
const_rate_interval = 0.05
const_lost_w_rate =  1 - const_rate_interval
const_benefit_w_rate =  1 + const_rate_interval

const_last_column = 31
const_row_max = 9999

def datetime_toString(dt):
    return dt.strftime("%Y-%m-%d")

def string_toDatetime(string):
    return dt.datetime.strptime(string, "%Y-%m-%d")

def string_toTimestamp(strTime):
    return dt.mktime(string_toDatetime(strTime).timetuple())

def timestamp_toString(stamp):
    return dt.strftime("%Y-%m-%d", time.localtime(stamp))

def datetime_toTimestamp(dateTim):
    return dt.mktime(dateTim.timetuple())

def day_string_plus_one(_day_input):
    date = string_toDatetime(_day_input) + dt.timedelta(days=1)
    return datetime_toString(date)

def get_hist_data_days(_stock_code, _day_start, _day_counts):
    df = ts.get_hist_data(_stock_code, start=_day_start, end=_day_start)
    if df is None:
        return df
    return df[['open', 'close', 'high', 'low']]

def search_data(stock_code, input_day, input_limit):
    count = 0
    a_store_data = [[]]
    try_times = 0
    while (count < input_limit):
        df = get_hist_data_days(stock_code, input_day, 1)

        if df is None or len(df) == 0:
            print 'got nothing:', 'stock:', stock_code, "day: ", input_day
            input_day = day_string_plus_one(input_day)
            try_times = try_times + 1
            if try_times == 10:
                return a_store_data
            continue
        
        close_price = df[u'close']
        open_price = df[u'open']
        high_price = df[u'high']
        low_price = df[u'low']

        a_store_data.append([open_price[0], close_price[0], high_price[0], low_price[0]])
        input_day = day_string_plus_one(input_day)
        count = count + 1
    return a_store_data

def write_to_cell(ws, row_index, column_index, value):
    print "write value: ", value, " to row=", row_index, "column=", column_index
    ws.cell(row=row_index, column=column_index).value = value

# when lost_rate < const_rate_interval, we need to sell it out
def is_needed_buy_in(ws, row_index, columns, price, count):
    print  __name__, ":", "input price is ", price
    mid_price = price
    for col_index in columns:
        get_price = ws.cell(row=row_index, column=col_index).value
        if get_price is None:
            print  __name__, ":", "get Price is None"
            break

        if mid_price > get_price:
            print  __name__, ":", "set mid_price to ", get_price
            mid_price = get_price
        
    if  mid_price <= (price * const_lost_w_rate):
        print  __name__, ":", "need 2 times buy in to update..."
        return True            
    return False

# once profit_rate > 5%, we need to sell it out
def is_needed_sell_out(ws, row_index, columns, price, count):
    print  __name__, ":", "input price is ", price
    mid_price = price
    for col_index in columns:
        get_price = ws.cell(row=row_index, column=col_index).value
        if get_price is None:
            print  __name__, ":", "get Price is None"
            break

        if (price * const_benefit_w_rate) <= get_price:
            print  __name__, ":", "sell price is ", (price * const_benefit_w_rate)
            #  is_sell , sell_price, total_count, total_amount
            return (True, (price * const_benefit_w_rate), count, (price * const_benefit_w_rate * count))
    
    print  __name__, ":", "will not sell in this check....."
    return (True, (price * const_benefit_w_rate), count, (price * const_benefit_w_rate * count))

# return the rate of profit
def to_sell_in_final_day_with_close_price(ws, row_index, columns, price):
    for col_index in columns:
        get_price = ws.cell(row=row_index, column=col_index).value
            
        if get_price is None:
            print "get Price is None"
            return (-1)
    print "sell price of the final day: ", get_price
    return ((get_price - price) / get_price) * 100
    
                
def calculate_from_excel():
    dirname, filename = os.path.split(os.path.abspath(__file__))
    read_excel_file_name = dirname + '\stock_s.xlsx'

    print read_excel_file_name
    wb = load_workbook(read_excel_file_name)

    sheets = wb.sheetnames
    sheet_first = sheets[0]
    #
    ws = wb[sheet_first]
    print "***"
    print sheet_first
    print ws.title
    print "^^^"
    
    rows = ws.rows

    input_limit = 5

    #print "stock_code: ",  "date: ",  "close: ",  "open: ", "high: ", "low: "

    row_index = 3   # start from row 3 to calculate

    while row_index < const_row_max:

        total_count = 0
        average_price = 0

        stock_code = ws.cell(row=row_index, column=2).value
        if stock_code is None:
            print "Have finish the calculation......"
            break

        print "Going to calculate code: ", stock_code

        # the open price of the first day
        average_price = ws.cell(row=row_index, column=12).value
        if average_price is None:
            row_index = row_index + 1
            continue
        total_count = calc_count(average_price, setup_base_line)
        total_amount = average_price * total_count
        
        print "buy in code: ", stock_code, ", price: ", average_price, ", total count: ", total_count, ", totoal amount: ", total_amount

        #is_sell = is_needed_sell_out(ws, row_index, columns_mid_1_day, average_price, total_count)
        is_sell, sell_price, total_count, total_amount = is_needed_sell_out(
            ws, row_index, columns_mid_1_day, average_price, total_count)
        if is_sell == True:
            print __name__, ":", stock_code, " is selled in the first day w ",
            print " average_price : ", average_price, " sell_price : ", sell_price,
            print " total_count : ", total_count, " total_amount : ", total_amount,
            print " profit_rate : ", const_rate_interval
            write_to_cell(ws, row_index, const_last_column,
                          const_rate_interval)
            row_index = row_index + 1
            continue

        if is_needed_buy_in(ws, row_index, columns_mid_1_day, average_price, total_count) == True:
            total_count = total_count * 3
            average_price = calc_price(average_price)
            total_amount = average_price * total_count

        print "after second day, code: ", stock_code, ", price: ", average_price, ", total count: ", total_count, ", totoal amount: ", total_amount

        is_sell, sell_price, total_count, total_amount = is_needed_sell_out(
            ws, row_index, columns_mid_1_day, average_price, total_count)
        if is_sell == True:
            print __name__, ":", stock_code, " is selled in the first day w ",
            print " average_price : ", average_price, " sell_price : ", sell_price,
            print " total_count : ", total_count, " total_amount : ", total_amount,
            print " profit_rate : ", const_rate_interval
            write_to_cell(ws, row_index, const_last_column,
                          const_rate_interval)
            row_index = row_index + 1
            continue

        if is_needed_buy_in(ws, row_index, columns_mid_2_day, average_price, total_count) == True:
            total_count = total_count * 3
            average_price = calc_price(average_price)
            total_amount = average_price * total_count

        print "after third day, code: ", stock_code, ", price: ", average_price, ", total count: ", total_count, ", totoal amount: ", total_amount

        print "after calculation, with the open price of the forth day......", " average_price: ", average_price, "total_amount: ", total_amount
        write_to_cell(ws, row_index, const_last_column, to_sell_in_final_day_with_close_price(
            ws, row_index, columns_last_day, average_price))

        row_index = row_index + 1

    print "save calculation result to ", const_last_column
    wb.save('stock_s.xlsx')
    return

def calc_price(price):  # standard_f 0.97
    return ((1 + 2 * const_lost_w_rate) * price / 3)

def calc_count(price, set_up_value):
    return (((set_up_value // price) // 100) * 100)
    
def main():
    calculate_from_excel()
    return

main()
