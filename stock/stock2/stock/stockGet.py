import sys
import os
import datetime as dt
import json
import tushare as ts
from openpyxl import load_workbook

#def printChinese(chinese_str):
#   type = sys.getfilesystemencoding()
#   print chinese_str.decode('UTF-8').encode(type)
#   return

#把datetime转成字符串
def datetime_toString(dt):
    return dt.strftime("%Y-%m-%d")

#把字符串转成datetime
def string_toDatetime(string):
    return dt.datetime.strptime(string, "%Y-%m-%d")

#把字符串转成时间戳形式
def string_toTimestamp(strTime):
    return dt.mktime(string_toDatetime(strTime).timetuple())

#把时间戳转成字符串形式
def timestamp_toString(stamp):
    return dt.strftime("%Y-%m-%d", time.localtime(stamp))

#把datetime类型转外时间戳形式
def datetime_toTimestamp(dateTim):
    return dt.mktime(dateTim.timetuple())

#将字符串e.g.'2018-07-06'转化为datetime 然后+1，在转回字符串'2018-07-07'
def day_string_plus_one(_day_input):
    date = string_toDatetime(_day_input) + dt.timedelta(days=1)
    return datetime_toString(date)

def get_hist_data_days(_stock_code, _day_start, _day_counts):
    df = ts.get_hist_data(_stock_code, start=_day_start, end=_day_start)
    if df is None:
        return df
    return df[['open', 'close', 'high', 'low']]

def search_data(stock_code, input_day, input_limit):
    count = 0
    a_store_data = [[]]
    try_times = 0
    while (count < input_limit):
        df = get_hist_data_days(stock_code, input_day, 1)

        if df is None or len(df) == 0:
            print 'got nothing:', 'stock:', stock_code, "day: ", input_day
            input_day = day_string_plus_one(input_day)
            try_times = try_times + 1
            if try_times == 10:
                return a_store_data
            continue
        
        close_price = df[u'close']
        open_price = df[u'open']
        high_price = df[u'high']
        low_price = df[u'low']
        #p_change_interval = df[u'p_change']
           
        # there is no data in that day, move to next day and have a try
        #if 0 == len(close_price) or 0 == len(open_price) or 0 == len(high_price) or 0 == len(low_price):
            #print "There is no data in day: ", input_day
        #    input_day = day_string_plus_one(input_day)
            #print "Try next day:", input_day
        #    continue
        a_store_data.append([open_price[0], close_price[0], high_price[0], low_price[0]])
        input_day = day_string_plus_one(input_day)
        count = count + 1
    return a_store_data

def main():
    #df = ts.get_hist_data('002415', start = '2018-07-06', end = '2018-07-10')
    #get_hist_data_days('002415', '2018-07-06', 1)
    #
    #df = ts.get_hist_data('000875')
    dirname, filename = os.path.split(os.path.abspath(__file__))
    #print dirname, filename
    read_excel_file_name = dirname + '\stock_s.xlsx'

    print read_excel_file_name

    wb = load_workbook(read_excel_file_name)

    # 获取所有表格(worksheet)的名字
    sheets = wb.sheetnames
    #print sheets
    # # 第一个表格的名称
    sheet_first = sheets[0]
    # # 获取特定的worksheet
    #
    ws = wb[sheet_first]
    rows = ws.rows

    print "***"
    print sheet_first
    print ws.title
    print "^^^"

    input_limit = 5

    print "stock_code: ",  "date: ",  "close: ",  "open: ", "high: ", "low: "
    # 迭代所有的行
    row_index = 1
    for row in rows:      
        line = [col.value for col in row]
        #print line
        if row_index == 1:
            row_index = row_index + 1
            continue
        if row_index == 2:
            row_index = row_index + 1
            continue

        stock_code = ws.cell(row=row_index, column=2).value

        if stock_code == None:
            break

        print "going to get data for stock: ", stock_code

        # go to get 
        get_day_from_table = ws.cell(row=row_index, column=3).value
        print get_day_from_table
        if get_day_from_table == None:
            print "break when day is None"
            break
        got_datas_one_stock = search_data(stock_code, datetime_toString(get_day_from_table), input_limit)
        if len(got_datas_one_stock) == 0:
            print "could not get data for stock:", stock_code, "move to next stock"
            row_index = row_index + 1
            
        item_index = 7
        for one_data in got_datas_one_stock:
            flag_item_index = item_index
            for item in one_data:
                ws.cell(row=row_index, column=item_index).value = item
                item_index = item_index + 1
                if item_index == (flag_item_index + 4):
                    item_index = item_index + 1
        
        # move to next row to search 
        row_index = row_index + 1
    print "goning to save data in to file"
    wb.save('stock_s.xlsx')
    return
 
main()
