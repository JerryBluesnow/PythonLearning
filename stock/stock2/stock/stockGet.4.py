import sys
import os
import datetime as dt
import json
import tushare as ts
from openpyxl import load_workbook

#def printChinese(chinese_str):
#   type = sys.getfilesystemencoding()
#   print chinese_str.decode('UTF-8').encode(type)
#   return

def datetime_toString(dt):
    return dt.strftime("%Y-%m-%d")

def string_toDatetime(string):
    return dt.datetime.strptime(string, "%Y-%m-%d")

def string_toTimestamp(strTime):
    return dt.mktime(string_toDatetime(strTime).timetuple())

def timestamp_toString(stamp):
    return dt.strftime("%Y-%m-%d", time.localtime(stamp))

def datetime_toTimestamp(dateTim):
    return dt.mktime(dateTim.timetuple())

def day_string_plus_one(_day_input):
    date = string_toDatetime(_day_input) + dt.timedelta(days=1)
    return datetime_toString(date)

def get_hist_data_days(_stock_code, _day_start, _day_counts):
    df = ts.get_hist_data(_stock_code, start=_day_start, end=_day_start)
    if df is None:
        return df
    return df[['open', 'close', 'high', 'low']]

def search_data(stock_code, input_day, input_limit):
    count = 0
    a_store_data = [[]]
    try_times = 0
    while (count < input_limit):
        df = get_hist_data_days(stock_code, input_day, 1)

        if df is None or len(df) == 0:
            print 'got nothing:', 'stock:', stock_code, "day: ", input_day
            input_day = day_string_plus_one(input_day)
            try_times = try_times + 1
            if try_times == 10:
                return a_store_data
            continue
        
        close_price = df[u'close']
        open_price = df[u'open']
        high_price = df[u'high']
        low_price = df[u'low']
        #p_change_interval = df[u'p_change']
           
        # there is no data in that day, move to next day and have a try
        #if 0 == len(close_price) or 0 == len(open_price) or 0 == len(high_price) or 0 == len(low_price):
            #print "There is no data in day: ", input_day
        #    input_day = day_string_plus_one(input_day)
            #print "Try next day:", input_day
        #    continue
        a_store_data.append([open_price[0], close_price[0], high_price[0], low_price[0]])
        input_day = day_string_plus_one(input_day)
        count = count + 1
    return a_store_data

def write_to_cell(ws, row_index, column_index, value):
    print "write value: ", value, " to row=", row_index, "column=", column_index
    ws.cell(row=row_index, column=column_index).value = value

def get_save_data_excel():
    #get file path and file name
    dirname, filename = os.path.split(os.path.abspath(__file__))

    read_excel_file_name = dirname + '\stock_s.xlsx'

    print read_excel_file_name

    wb = load_workbook(read_excel_file_name)

    sheets = wb.sheetnames
    #print sheets
    sheet_first = sheets[0]
    #
    ws = wb[sheet_first]
    rows = ws.rows

    print "***"
    print sheet_first
    print ws.title
    print "^^^"

    input_limit = 5

    print "stock_code: ",  "date: ",  "close: ",  "open: ", "high: ", "low: "
    row_index = 1
    for row in rows:      
        line = [col.value for col in row]
        #print line
        if row_index == 1:
            row_index = row_index + 1
            continue
        if row_index == 2:
            row_index = row_index + 1
            continue

        stock_code = ws.cell(row=row_index, column=2).value

        if stock_code == None:
            break

        print "going to get data for stock: ", stock_code

        # go to get 
        get_day_from_table = ws.cell(row=row_index, column=3).value
        print get_day_from_table
        if get_day_from_table == None:
            print "break when day is None"
            break
        got_datas_one_stock = search_data(stock_code, datetime_toString(get_day_from_table), input_limit)
        if len(got_datas_one_stock) == 0:
            print "could not get data for stock:", stock_code, "move to next stock"
            row_index = row_index + 1
            
        item_index = 7
        for one_data in got_datas_one_stock:
            flag_item_index = item_index
            for item in one_data:
                ws.cell(row=row_index, column=item_index).value = item
                item_index = item_index + 1
                if item_index == (flag_item_index + 4):
                    item_index = item_index + 1
        
        # move to next row to search 
        row_index = row_index + 1
    print "goning to save data in to file"
    wb.save('stock_s.xlsx')
    return

columns_first_day = [12, 13, 14, 15]
columns_mid_1_day = [17, 18, 19, 20]
columns_mid_2_day = [22, 23, 24, 25]
columns_last_day = [27, 28, 29, 30]
setup_base_line = 10000
const_standard_rate = 0.97
const_last_column = 31
const_row_max = 9999

def is_need_buy_in(ws, row_index, columns, price, count):
    print "input price is ", price
    mid_price = price
    for col_index in columns:
        get_price = ws.cell(row=row_index, column=col_index).value
        if get_price is None:
            print "get Price is None"
            break

        if mid_price > get_price:
            print "set mid_price to ", get_price
            mid_price = get_price
        
    if  mid_price <= (price * const_standard_rate):
        print "need 2 times buy in to update..."
        return True            
    return False

# return the rate of profit
def to_sell_in_final_day_with_close_price(ws, row_index, columns, price):
    for col_index in columns:
        get_price = ws.cell(row=row_index, column=col_index).value
            
        if get_price is None:
            print "get Price is None"
            return (-1)
    print "sell price of the final day: ", get_price
    return ((get_price - price) / get_price) * 100
    
                
def calculate_from_excel():
    dirname, filename = os.path.split(os.path.abspath(__file__))
    read_excel_file_name = dirname + '\stock_s.xlsx'

    print read_excel_file_name
    wb = load_workbook(read_excel_file_name)

    sheets = wb.sheetnames
    sheet_first = sheets[0]
    #
    ws = wb[sheet_first]
    print "***"
    print sheet_first
    print ws.title
    print "^^^"
    
    rows = ws.rows

    input_limit = 5

    #print "stock_code: ",  "date: ",  "close: ",  "open: ", "high: ", "low: "

    row_index = 3   # start from row 3 to calculate

    while row_index < const_row_max:

        total_count = 0
        average_price = 0

        stock_code = ws.cell(row=row_index, column=2).value
        if stock_code is None:
            print "Have finish the calculation......"
            break

        print "Going to calculate code: ", stock_code

        # the open price of the first day
        average_price = ws.cell(row=row_index, column=12).value
        if average_price is None:
            row_index = row_index + 1
            continue
        total_count = calc_count(average_price, setup_base_line)
        total_amount = average_price * total_count
        
        print "buy in code: ", stock_code, ", price: ", average_price, ", total count: ", total_count, ", totoal amount: ", total_amount

        if is_need_buy_in(ws, row_index, columns_mid_1_day, average_price, total_count) == True:
            total_count = total_count * 3
            average_price = calc_price(average_price)
            total_amount = average_price * total_count
        
        print "after second day, code: ", stock_code, ", price: ", average_price, ", total count: ", total_count, ", totoal amount: ", total_amount
        
        if is_need_buy_in(ws, row_index, columns_mid_2_day, average_price, total_count) == True:
            total_count = total_count * 3
            average_price = calc_price(average_price)
            total_amount = average_price * total_count
        
        print "after third day, code: ", stock_code, ", price: ", average_price, ", total count: ", total_count, ", totoal amount: ", total_amount
        
        print "after calculation, with the open price of the forth day......", " average_price: ", average_price, "total_amount: ", total_amount
        write_to_cell(ws, row_index, const_last_column, to_sell_in_final_day_with_close_price(ws, row_index, columns_last_day, average_price)) 

        row_index = row_index + 1
    
    print "save calculation result to ", const_last_column
    wb.save('stock_s.xlsx')
    return

def calc_price(price):  # standard_f 0.97
    return ((1 + 2 * const_standard_rate) * price / 3)

def calc_count(price, set_up_value):
    return (((set_up_value // price) // 100) * 100)
    
def main():
    calculate_from_excel()
    return

main()
