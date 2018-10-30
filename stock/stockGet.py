#-*- coding: utf-8 -*-

import sys
import datetime as dt
import json
import tushare as ts
from openpyxl import load_workbook

#from openpyxl.reader.excel import load_workbook

##wb = load_workbook(filename=r'D:/PythonWorkplace/PythonLearning/stock/stock_s_base.xlsx')

#print "Worksheet range(s):", wb.get_named_ranges()
#print "Worksheet name(s):", wb.get_sheet_names()

def printChinese(chinese_str):
    type = sys.getfilesystemencoding()
    print chinese_str.decode('UTF-8').encode(type)
    return

#saved_filename = 'stock.xlsx'

#df = ts.get_hist_data('002415', start='2018-05-28', end='2018-05-30')

#df.to_excel(saved_filename)

#把datetime转成字符串
def datetime_toString(dt):
    return dt.strftime("%Y-%m-%d")

#把字符串转成datetime
def string_toDatetime(string):
    return dt.datetime.strptime(string, "%Y-%m-%d")

#把字符串转成时间戳形式
def string_toTimestamp(strTime):
    return dt.mktime(string_toDatetime(strTime).timetuple())

#把时间戳转成字符串形式
def timestamp_toString(stamp):
    return dt.strftime("%Y-%m-%d", tiem.localtime(stamp))

#把datetime类型转外时间戳形式
def datetime_toTimestamp(dateTim):
    return dt.mktime(dateTim.timetuple())

#将字符串e.g.'2018-07-06'转化为datetime 然后+1，在转回字符串'2018-07-07'
def day_string_plus_one(_day_input):
    date = string_toDatetime(_day_input) + dt.timedelta(days=1)
    return datetime_toString(date)

def get_hist_data_days(_stock_code, _day_start, _day_counts):
    df = ts.get_hist_data(_stock_code, start = _day_start, end = _day_start)
    return df[['open', 'close', 'high', 'low']]

def search_data(stock_code, input_day, input_limit):
    count = 0
    while (count < input_limit):
        df = get_hist_data_days(stock_code, input_day, 1)

        close_price = df[u'close']
        open_price = df[u'open']
        high_price = df[u'high']
        low_price = df[u'low']
           
        # there is no data in that day, move to next day and have a try
        if 0 == len(close_price):
            print "There is no data in day: ", input_day
            input_day = day_string_plus_one(input_day)
            print "Try next day:", input_day
            continue
        #rint tuple(df)
        #print df.open, df.close, df.high, df.low

        #设定数据位置（从第3行，第6列开始插入数据）
        #df.to_csv('D:/PythonWorkplace/PythonLearning/stock/1.xlsx')
        #df.to_excel('D:/PythonWorkplace/PythonLearning/stock/stock_s.xlsx', startrow=2,startcol=5)
        count = count + 1
        
        print "stock_code: ", stock_code, "close: ", close_price[0], "open: ", open_price[0], "high: ", high_price[0], "low: ", low_price[0]
        input_day = day_string_plus_one(input_day)
        #print count
    pass

def main():
    #df = ts.get_hist_data('002415', start = '2018-07-06', end = '2018-07-10')
    #get_hist_data_days('002415', '2018-07-06', 1)
    #
    #df = ts.get_hist_data('000875')
    read_excel_file_name = 'D:/PythonWorkplace/PythonLearning/stock/stock_s.xlsx'
    #df.to_excel('D:/PythonWorkplace/PythonLearning/stock/1.xlsx')
    wb = load_workbook(read_excel_file_name)
    #wb = load_workbook(filename=ExcelFullName)

    # 获取当前活跃的worksheet,默认就是第一个worksheet
    #ws = wb.active
    # 当然也可以使用下面的方法
    # 获取所有表格(worksheet)的名字
    sheets = wb.get_sheet_names()
    print sheets
    # # 第一个表格的名称
    sheet_first = sheets[0]
    # # 获取特定的worksheet
    #
    ws = wb[sheet_first]
    print "***"
    print sheet_first
    print ws.title
    print "^^^"
    # 获取表格所有行和列，两者都是可迭代的

    rows = ws.rows
    print rows

    input_limit = 4

    # 迭代所有的行
    row_index = 1 
    for row in rows:      
        line = [col.value for col in row]
        #print line
        if row_index == 1:
            row_index = row_index + 1
            continue
        if row_index == 2:
            row_index = row_index + 1
            continue
        stock_code = ws.cell(row=row_index, column=2).value
        #print datetime_toString(ws.cell(row=row_index, column=3).value)
        input_day = datetime_toString(ws.cell(row=row_index, column=3).value)
        search_data(stock_code, input_day, input_limit)
        row_index = row_index + 1

    return
 
main()
