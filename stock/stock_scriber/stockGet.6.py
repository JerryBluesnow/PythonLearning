#!/usr/bin/python
# -*- coding: utf-8 -*-
import sys
import os
import datetime as dt
import json
import tushare as ts
from openpyxl import load_workbook
from optparse import OptionParser

_DEBUG_LOG = 0

OUTPUT_LIMIT_MAX        = 5
ROW_START_INDEX         = 3
COLUMN_FIRST_OUTPUT     = 7
COLUMN_COUNT_OUTPUT     = 4
COLUMN_COUNT_PER_ITEM   = 6
COLUMN_OF_START_DAY     = 3
COLUMN_OF_STOCK_CODE    = 2
COLUMN_OF_STOCK_NAME    = 1
CONST_2_SEARCH_ROW_MAX  = 5

def datetime_toString(dt):
    return dt.strftime("%Y-%m-%d")

def string_toDatetime(string):
    return dt.datetime.strptime(string, "%Y-%m-%d")

def string_toTimestamp(strTime):
    return dt.mktime(string_toDatetime(strTime).timetuple())

def timestamp_toString(stamp):
    return dt.strftime("%Y/%m-%d", time.localtime(stamp))

def datetime_toTimestamp(dateTim):
    return dt.mktime(dateTim.timetuple())

def day_string_plus_one(_day_input):
    date = string_toDatetime(_day_input) + dt.timedelta(days=1)
    return datetime_toString(date)

def get_hist_data_days(_stock_code, _day_start, _day_counts):
    df = ts.get_hist_data(_stock_code, start=_day_start, end=_day_start)
    if df is None:
        return df
    return df[['open', 'close', 'high', 'low']]

def search_data(stock_code, input_day, OUTPUT_LIMIT_MAX):
    '''
    stock_code          - python string
    input_day           - must in excel in date format
    OUTPUT_LIMIT_MAX    - how many days will be query
    '''
    count = 0
    a_store_data = []
    try_times = 0

    #day_2_search = input_day.replace("/", "-")
    day_2_search = datetime_toString(input_day)

    print '++++++++++++++++++++++++++++++++++++++++++++++++++++++'
    print 'stock code information collection status:'
    while (count < OUTPUT_LIMIT_MAX):
        df = get_hist_data_days(stock_code, day_2_search, 1)

        if df is None or len(df) == 0:
            print 'got nothing of day: ', day_2_search
            day_2_search = day_string_plus_one(day_2_search)
            try_times = try_times + 1
            if try_times == 10:
                return a_store_data
            continue
        
        print 'got data of the day:', day_2_search, 
        close_price = round(df[u'close'], 2)
        open_price = round(df[u'open'], 2)
        high_price = round(df[u'high'], 2)
        low_price = round(df[u'low'], 2)

        one_item = [open_price, close_price, high_price, low_price]
        a_store_data.append(one_item)
        print one_item
        day_2_search = day_string_plus_one(day_2_search)
        count = count + 1
    print '++++++++++++++++++++++++++++++++++++++++++++++++++++++'
    return a_store_data

def write_to_cell(ws, row_index, column_index, value):
    print "write value: ", value, " to row=", row_index, "column=", column_index
    ws.cell(row=row_index, column=column_index).value = value


def search_and_write_stock_data_2_excel(_data_file):
    #dirname, filename = os.path.split(os.path.abspath(__file__))
    #read_excel_file_name = dirname + '\stock_s.xlsx'
    
    read_excel_file_name = _data_file

    print read_excel_file_name
    wb = load_workbook(read_excel_file_name)

    sheets = wb.sheetnames
    sheet_first = sheets[0]
    #
    ws = wb[sheet_first]
    print "***"
    print sheet_first
    print ws.title
    print ws.max_row
    print "^^^"
    
    rows = ws.rows

    row_index = ROW_START_INDEX   # start from row 3 to calculate

    while row_index < ws.max_row:
        # get stock_code from column=2
        stock_code = ws.cell(row=row_index, column=COLUMN_OF_STOCK_CODE).value
        if stock_code is None:
            print "Have finish the calculation......"
            break

        # the open price of the first day
        search_start_day = ws.cell(row=row_index, column=COLUMN_OF_START_DAY).value
        
        print 'stock_code:',stock_code, ", search_start_day:", search_start_day
        returned_data_collection = search_data(stock_code, search_start_day, OUTPUT_LIMIT_MAX)
        
        if returned_data_collection is None or len(returned_data_collection) == 0:
            row_index = row_index + 1
            continue
        column_to_write = COLUMN_FIRST_OUTPUT
        for item in returned_data_collection:
            sub_index = 0
            for sub_item in item:
                write_to_cell(ws, row_index, column_to_write, item[sub_index])
                column_to_write = column_to_write + 1
                sub_index = sub_index + 1
            column_to_write = column_to_write + (COLUMN_COUNT_PER_ITEM - COLUMN_COUNT_OUTPUT)

        row_index = row_index + 1

    print "save calculation result to "
    wb.save('stock_s.xlsx')
    return

def parse_stock_code_from_xml(_code_file):
    stock_dict = {}
    with open(_code_file,'r') as f:
        lines = f.readlines()
        for line in lines:
            if line is None or len(line) == 0:
                continue
            line_after_replace = line.replace('<li>','')
            line_after_replace = line_after_replace.replace('</a></li>','')
            #print line_after_replace
            right_branket_pos = line_after_replace.rfind(')')
            left_branket_pos = line_after_replace.find('(') 
            stock_name_pos = line_after_replace.find('>') + 1
            stock_code = line_after_replace[left_branket_pos+1:right_branket_pos]
            stock_name = line_after_replace[stock_name_pos:left_branket_pos]
            stock_dict[stock_name.decode('gbk').encode('utf-8')] = stock_code
            #print stock_name, stock_code
    if _DEBUG_LOG == 1:
        print stock_dict
    return stock_dict

def get_stock_code_from_dict_by_name(_stock_name, _stock_dict):
    #print _stock_name.decode("string-escape")
    #print _stock_name
    try:
        return _stock_dict[_stock_name.encode('utf-8')]
    except:
        print "Couldn't get stock code from dict..."
        return None

def get_stock_code_from_dict_by_name_all_to_excell(_stock_dict, _data_file):
    if _stock_dict is None or len(_stock_dict) == 0:
        print "No data in _stock_dict, exit..."
        return
    if os.path.isfile(_data_file) != True:
        print "input _data_file is not a file:", _data_file
        return 

    #dirname, filename = os.path.split(os.path.abspath(__file__))
    #read_excel_file_name = dirname + '\stock_s.xlsx'
    read_excel_file_name = _data_file
    print read_excel_file_name
    wb = load_workbook(read_excel_file_name)
    
    sheets = wb.sheetnames
    sheet_first = sheets[0]
    #
    ws = wb[sheet_first]
    print "***"
    print sheet_first
    print ws.title
    print ws.max_row
    print "^^^"
    
    rows = ws.rows
    row_index = ROW_START_INDEX   # start from row 3 to calculate

    while row_index <= ws.max_row:
        # get stock_code from column=2
        stock_name = ws.cell(row=row_index, column=COLUMN_OF_STOCK_NAME).value
        if stock_name is None:
            break
        
        stock_code = get_stock_code_from_dict_by_name(stock_name, _stock_dict)
        if stock_code is not None:
            #print stock_name.encode('utf-8'), stock_code
            print stock_name, stock_code
            write_to_cell(ws, row_index, COLUMN_OF_STOCK_CODE, stock_code)
        row_index = row_index + 1

    wb.save('stock_s.xlsx')
    return

def main():
    '''All parameters are mandatory, please make sure they are used'''
    usage = "usage: %prog [options] arg"
    parser = OptionParser(usage)
    parser.add_option("-c", "--code", dest="code",
                      help="get stock code by input.xml")
    parser.add_option("-d", "--data", dest="data",
                      help="data files as input and output")
    parser.add_option("-v", "--verbose",
                      action="store_true", dest="verbose")
    parser.add_option("-q", "--quiet",
                      action="store_false", dest="verbose")
    (options, args) = parser.parse_args()
    #reload(sys)
    #sys.setdefaultencoding("string-escape")
    #search_data('600547','2018/09/19',5)
    #search_and_write_stock_data_2_excel()

    if os.path.isfile(options.data) != True:
        print "no input data file, exit...."
        return

    if os.path.isfile(options.code) != True:
        print "++++++++++++++++++START++++++++++++++++++++++++++++"
        print "+-going to parse stock code from file:", options.code
        stock_dict = parse_stock_code_from_xml(options.code)
        if stock_dict is None or len(stock_dict) == 0:
            return
        get_stock_code_from_dict_by_name_all_to_excell(stock_dict, options.data)
        print "+Complete..."
        return
        print "+++++++++++++++++++END+++++++++++++++++++++++++++++"

    print "++++++++++++++++++START++++++++++++++++++++++++++++"
    search_and_write_stock_data_2_excel(options.data)
    print "+++++++++++++++++++END+++++++++++++++++++++++++++++"
    
    return

if __name__ == "__main__":
    main()
